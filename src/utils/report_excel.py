"""Excel report export for squat video pipeline artifacts."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any


def export_video_report(run_dir: str | Path, output_path: str | Path | None = None) -> Path:
    """Create an Excel workbook from an existing video pipeline run directory."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export.") from exc

    run_path = Path(run_dir)
    summary_path = run_path / "video_summary.json"
    metrics_path = run_path / "video_frames_metrics.csv"
    if not summary_path.is_file():
        raise FileNotFoundError(f"Video summary not found: {summary_path}")
    if not metrics_path.is_file():
        raise FileNotFoundError(f"Frame metrics CSV not found: {metrics_path}")

    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    rows = _read_csv_rows(metrics_path)
    report_path = Path(output_path) if output_path else _default_report_path(summary)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary_sheet(summary_sheet, summary)
    _write_frame_metrics_sheet(workbook.create_sheet("Frame metrics"), rows)
    _write_warnings_sheet(workbook.create_sheet("Warnings"), summary, rows)
    _write_artifacts_sheet(workbook.create_sheet("Artifacts"), summary)
    workbook.save(report_path)
    return report_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export an Excel report from a video pipeline run directory.")
    parser.add_argument("--run-dir", required=True, help="Path to outputs/predictions/<run_id>.")
    parser.add_argument("--output", default=None, help="Optional XLSX output path.")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    try:
        report_path = export_video_report(args.run_dir, args.output)
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        parser.exit(status=1, message=f"Error: {exc}\n")
    print(str(report_path))


def _write_summary_sheet(sheet: Any, summary: dict[str, Any]) -> None:
    sheet.append(("field", "value"))
    fields = (
        ("run_id", summary.get("run_id")),
        ("input video", summary.get("input_path")),
        ("exercise type", summary.get("exercise_type", "squat")),
        ("model", summary.get("model_name")),
        ("backend", summary.get("backend")),
        ("bbox strategy", summary.get("bbox_strategy")),
        ("frame stride", summary.get("frame_stride")),
        ("processed frames", summary.get("processed_frame_count")),
        ("valid keypoint frame ratio", summary.get("valid_keypoint_frame_ratio")),
        ("usable knee frame ratio", summary.get("usable_knee_frame_ratio")),
        ("reliable frame ratio", summary.get("reliable_frame_ratio")),
        ("min selected knee angle", summary.get("min_selected_knee_angle")),
        ("max selected knee angle", summary.get("max_selected_knee_angle")),
        ("mean torso lean angle", summary.get("mean_torso_lean_angle")),
        ("depth status", summary.get("depth_status")),
        ("warning frame count", summary.get("warning_frame_count")),
        ("brief conclusion", _brief_conclusion(summary)),
    )
    for field, value in fields:
        sheet.append((field, value))


def _write_frame_metrics_sheet(sheet: Any, rows: list[dict[str, str]]) -> None:
    if not rows:
        sheet.append(("none",))
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([_coerce_cell(row.get(header, "")) for header in headers])


def _write_warnings_sheet(sheet: Any, summary: dict[str, Any], rows: list[dict[str, str]]) -> None:
    counter: Counter[str] = Counter()
    for row in rows:
        for warning in _split_warnings(row.get("warnings", "")):
            counter[warning] += 1
    if not counter:
        for warning, count in dict(summary.get("dominant_warnings", {})).items():
            counter[str(warning)] = int(count)
    sheet.append(("warning", "frame_count"))
    if not counter:
        sheet.append(("none", 0))
        return
    for warning, count in counter.most_common():
        sheet.append((warning, count))


def _write_artifacts_sheet(sheet: Any, summary: dict[str, Any]) -> None:
    sheet.append(("artifact", "path"))
    artifacts = dict(summary.get("artifacts", {}))
    ordered_keys = (
        "summary",
        "frame_metrics_csv",
        "keypoints_json",
        "sampled_frames_dir",
        "annotated_video",
        "annotated_video_browser",
        "excel_report",
        "history_jsonl",
    )
    wrote = False
    for key in ordered_keys:
        value = artifacts.get(key)
        if value:
            sheet.append((key, value))
            wrote = True
    if not wrote:
        sheet.append(("none", ""))


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as file:
        return list(csv.DictReader(file))


def _default_report_path(summary: dict[str, Any]) -> Path:
    run_id = summary.get("run_id") or "unknown_run"
    return Path("outputs/reports") / f"video_report_{run_id}.xlsx"


def _brief_conclusion(summary: dict[str, Any]) -> str:
    if summary.get("status") != "ok":
        return "Run did not complete successfully; review the summary error message."
    if summary.get("depth_status") == "sufficient" and int(summary.get("warning_frame_count") or 0) == 0:
        return "Squat analysis completed with sufficient depth and no frame warnings."
    if summary.get("depth_status") == "insufficient":
        return "Squat analysis completed, but the run-level depth rule marked depth as insufficient."
    return "Squat analysis completed; review warnings and frame metrics."


def _split_warnings(value: str) -> list[str]:
    if not value:
        return []
    return [item for item in (part.strip() for part in value.split(";")) if item]


def _coerce_cell(value: Any) -> Any:
    if value == "":
        return None
    if value in {"True", "False"}:
        return value == "True"
    try:
        integer = int(value)
    except (TypeError, ValueError):
        integer = None
    if integer is not None and str(integer) == str(value):
        return integer
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


if __name__ == "__main__":
    main()
