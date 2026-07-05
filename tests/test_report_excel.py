from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from src.utils.report_excel import build_parser, export_video_report


def test_excel_report_created_from_mocked_summary_and_csv(tmp_path: Path) -> None:
    run_dir = _write_mock_run(tmp_path, warnings="")

    report_path = export_video_report(run_dir, tmp_path / "report.xlsx")

    assert report_path.is_file()
    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Summary", "Frame metrics", "Warnings", "Artifacts"]
    assert workbook["Summary"]["B2"].value == "run-1"
    assert workbook["Frame metrics"]["A2"].value == 0
    assert workbook["Frame metrics"]["B2"].value == 170.0
    assert workbook["Warnings"]["A2"].value == "none"


def test_excel_warnings_sheet_aggregates_frame_warnings(tmp_path: Path) -> None:
    run_dir = _write_mock_run(tmp_path, warnings="low_confidence;angle_tentative")

    report_path = export_video_report(run_dir, tmp_path / "custom.xlsx")

    workbook = load_workbook(report_path)
    warnings_sheet = workbook["Warnings"]
    warnings = {warnings_sheet["A2"].value, warnings_sheet["A3"].value}
    assert warnings == {"low_confidence", "angle_tentative"}


def test_report_excel_cli_has_run_dir_argument() -> None:
    parser = build_parser()
    option_strings = {option for action in parser._actions for option in action.option_strings}

    assert "--run-dir" in option_strings
    assert "--output" in option_strings


def _write_mock_run(tmp_path: Path, warnings: str) -> Path:
    run_dir = tmp_path / "predictions" / "run-1"
    run_dir.mkdir(parents=True)
    summary = {
        "run_id": "run-1",
        "exercise_type": "squat",
        "input_path": "data/samples/squat_sample.mp4",
        "model_name": "RTMPose-S",
        "backend": "mmpose",
        "bbox_strategy": "whole_frame",
        "frame_stride": 5,
        "processed_frame_count": 1,
        "valid_keypoint_frame_ratio": 1.0,
        "usable_knee_frame_ratio": 1.0,
        "reliable_frame_ratio": 1.0,
        "min_selected_knee_angle": 170.0,
        "max_selected_knee_angle": 170.0,
        "mean_torso_lean_angle": 10.0,
        "depth_status": "insufficient",
        "warning_frame_count": 0 if not warnings else 1,
        "dominant_warnings": {},
        "status": "ok",
        "artifacts": {
            "summary": str(run_dir / "video_summary.json"),
            "frame_metrics_csv": str(run_dir / "video_frames_metrics.csv"),
            "keypoints_json": str(run_dir / "video_keypoints.json"),
            "sampled_frames_dir": str(run_dir / "sampled_frames"),
            "annotated_video": str(run_dir / "annotated_video.mp4"),
        },
    }
    (run_dir / "video_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    with (run_dir / "video_frames_metrics.csv").open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["frame_index", "selected_knee_angle", "warnings"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "frame_index": 0,
                "selected_knee_angle": 170.0,
                "warnings": warnings,
            }
        )
    return run_dir
